import logging, discord, datetime, os, openpyxl, re, pendulum, time
logger = logging.getLogger(__name__)

from discord.ext import commands
from discord import app_commands
from discord.ext.commands import has_any_role
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile

from modules.core.resources import Resources

from typing import Literal, Optional

class Day:
	ALL = 'all'
	SAT = 'saturday'
	WED = 'wednesday'

class Prosepect:
	NEXT = 'next'
	FUTURE = 'future'
	ALL = 'all'


class AnimeClub(commands.Cog):

	def __init__(self, bot):
		self.bot = bot

	def is_anime_club_server(ctx):
		return ctx.guild.id in [254864526069989377, 259896980308754432]

	async def cog_command_error(self, ctx, err):
		logger.exception("Error during Anime Club command.")
		try:
			await ctx.send('error!', file=discord.File(os.getcwd() + '/assets/lain_err_sm.png'))
		except:
			pass

	@app_commands.command(name="sc")
	@app_commands.describe(
		day='which day to check (defaults to all)',
		when='can show all, future, or just the next upcoming meeting (defaults to next)'
	)
	@app_commands.guilds(254864526069989377) # 259896980308754432
	async def s_sc(self, interaction, day: Optional[Literal[Day.SAT, Day.WED]] = Day.ALL, when: Optional[Literal[Prosepect.ALL, Prosepect.NEXT, Prosepect.FUTURE]] = Prosepect.NEXT):
		"""view club schedule"""
		if when == Prosepect.NEXT:
			await self.show_shcedule(interaction.response.send_message, sat=day == Day.ALL or day == Day.SAT, wed=day == Day.ALL or day == Day.WED)
		else:
			await interaction.response.defer()
			if day == Day.ALL or day == Day.SAT:
				await self.show_all_sat(interaction.followup.send, only_future=when == Prosepect.FUTURE)
			if day == Day.ALL or day == Day.WED:
				await self.show_all_wed(interaction.followup.send, only_future=when == Prosepect.FUTURE)

	@commands.group(aliases=['sced', 'sc'])
	@commands.check(is_anime_club_server)
	async def schedule(self, ctx):
		# await ctx.trigger_typing()
		if ctx.invoked_subcommand is None:
			if ctx.message.content in ['>sc', '>sced']:
				await self.show_shcedule(ctx.send, wed=True, sat=True)
			else:
				await ctx.send('Bad usage. Try `>sc` `>sc wed` or `>sc wed future`. Can replace wed with sat as well')

	@schedule.group(aliases=['sat', 'SAT', 'Sat', 'Saturday'])
	async def saturday(self, ctx):
		if ctx.invoked_subcommand is None:
			await self.show_shcedule(ctx.send, sat=True)

	@schedule.group(aliases=['wed', 'WED', 'Wed', 'Wednesday'])
	async def wednesday(self, ctx):
		if ctx.invoked_subcommand is None:
			await self.show_shcedule(ctx.send, wed=True)

	@schedule.command(name='all')
	async def all_both(self, ctx):
		await self.show_all_wed(ctx.send)
		await self.show_all_sat(ctx.send)

	@saturday.command(name='all')
	async def all_sat(self,ctx):
		await self.show_all_sat(ctx.send)

	@wednesday.command(name='all')
	async def all_wed(self, ctx):
		await self.show_all_wed(ctx.send)

	@schedule.command(name='future')
	async def future_both(self, ctx):
		await self.show_all_wed(ctx.send, only_future=True)
		await self.show_all_sat(ctx.send, only_future=True)

	@saturday.command(name='future')
	async def future_sat(self,ctx):
		await self.show_all_sat(ctx.send, only_future=True)

	@wednesday.command(name='future')
	async def future_wed(self, ctx):
		await self.show_all_wed(ctx.send, only_future=True)

	@app_commands.command(name="set-schedule")
	@app_commands.guilds(254864526069989377) # 259896980308754432
	@app_commands.checks.has_any_role(494979840470941712, 259557922726608896) # admin, executive council 260093344858767361
	async def s_set(self, interaction, day: Literal[Day.SAT, Day.WED], schedule: discord.Attachment):
		"""set club schedule (admin)"""
		if day == Day.SAT:
			k = 'Saturday'
			n = 6
		else:
			k = 'Wednesday'
			n = 7
		
		with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
			await schedule.save(tmp.name)
			try:
				sched = extract_schedule(tmp.name, n)
			except Exception as e:
				await interaction.response.send_message(f"Failed to extract schedule from file.\n{e.message}")
				return
			res = await Resources.storage_col.update_one({'id': 'sched_v2'}, {'$set': {k: sched}})
			if not res:
				await interaction.response.send_message(f"Error setting {k} schedule!")
			else:
				await interaction.response.send_message(f"{k} schedules have been updated!")

	async def show_shcedule(self, send, wed=False, sat=False):
		if not wed and not sat:
			return
		else:
			embed=discord.Embed(description="Club Schedule", color=0xd31f28)
			embed.set_thumbnail(url="https://files.catbox.moe/9dsqp5.png")
			data = await Resources.storage_col.find_one({'id': 'sched_v2'})

			if sat:
				nxt_sat = next_day(day=pendulum.SATURDAY)
				lines = saturday_lines(data['Saturday'].get(nxt_sat.format('YYYY-MM-DD HH:mm:ss')))
				if lines:
					embed.add_field(name=f"Saturday ({nxt_sat.month}/{nxt_sat.day})", value='\n'.join(lines), inline=False)
				else:
					embed.add_field(name=f"Saturday ({nxt_sat.month}/{nxt_sat.day})", value="*none*", inline=False)
			if wed:
				nxt_wed = next_day(day=pendulum.WEDNESDAY)
				lines = wednesday_lines(data['Wednesday'].get(nxt_wed.format('YYYY-MM-DD HH:mm:ss')))

				if lines:
					embed.add_field(name=f"Wednesday ({nxt_wed.month}/{nxt_wed.day})", value='\n'.join(lines), inline=True)
				else:
					embed.add_field(name=f"Wednesday ({nxt_wed.month}/{nxt_wed.day})", value="*none*", inline=False)

			await send(embed=embed)

	async def show_all_wed(self, send, only_future=False):
		embed=discord.Embed(description="Wednesday Schedules", color=0xd31f28)
		embed.set_thumbnail(url="https://files.catbox.moe/9dsqp5.png")

		data = await Resources.storage_col.find_one({'id': 'sched_v2'})
		if 'Wednesday' not in data:
			embed.add_field(name='Unavailable', value='could not find Wednesday data')
		else:
			for meeting in data['Wednesday']:
				date = pendulum.parse(meeting, tz=Resources.timezone_str)
				if only_future and date < next_day(day=pendulum.WEDNESDAY):
					continue
				lines = wednesday_lines(data['Wednesday'][meeting])
				if lines:
					embed.add_field(name=f"{date.month}/{date.day}", value='\n'.join(lines), inline=False)

		await send(embed=embed)

	async def show_all_sat(self, send, only_future=False):
		embed=discord.Embed(description="Saturday Schedules", color=0xd31f28)
		embed.set_thumbnail(url="https://files.catbox.moe/9dsqp5.png")

		data = await Resources.storage_col.find_one({'id': 'sched_v2'})
		if 'Saturday' not in data:
			embed.add_field(name='Unavailable', value='could not find Saturday data')
		else:
			for meeting in data['Saturday']:
				date = pendulum.parse(meeting, tz=Resources.timezone_str)
				if only_future and date < next_day(day=pendulum.SATURDAY):
					continue
				lines = saturday_lines(data['Saturday'][meeting])
				if lines:
					embed.add_field(name=f"{date.month}/{date.day}", value='\n'.join(lines), inline=False)

		await send(embed=embed)

def parse_title(title):
    # returns array of [title, episode]. if no episode, it will be None

    title = title.strip()

    ep = re.search('\d+$', title)

    if ep:
        return [re.sub('\d+$', '', title), ep.group()]
    else:
        ep = re.search('\(\d+\)$', title)
        if ep:
            return [re.sub('\(\d+\)$', '', title), ep.group()[1:-1]]
        else:
            return [title, None]

def wednesday_lines(data):
	return saturday_lines(data)
	lines = []
	if not data:
		return lines

	for showtime in data:
		try:
			if showtime['title'].lower().startswith('$break$'):
				lines.append(f"🚫 {showtime['title'][len('$break$'):].strip()}")
				break
		except:
			pass
			
		if showtime['title']:
			lines.append(f"{showtime['start']}-{showtime['end']}: {showtime['title']}")
	return lines

def saturday_lines(data):
	lines = []
	if not data:
		return lines

    # title: [start ep, end ep, time start]
	shows = {} # python 3.6+ dicts keep insertion order 🎉
	for showtime in data:
		if not showtime['title']:
			continue

		try:
			if showtime['title'].lower() == 'craptacular':
				lines.append('🎉💩 Craptacular 💩🎉')
				break
		except:
			pass

		try:
			if showtime['title'].lower().startswith('$break$'):
				lines.append(f"🚫 {showtime['title'][len('$break$'):].strip()}")
				break
		except:
			pass

		title = parse_title(showtime['title'])

		if title[0] in shows:
			shows[title[0]][1] = title[1]
		else:
			shows[title[0]] = [title[1], title[1], showtime['start']]

	for show in shows:
		if shows[show][0]:
			if shows[show][0] == shows[show][1]:
				lines.append(f"{shows[show][2]}pm: {show} ({shows[show][0]})")
			else:
				lines.append(f"{shows[show][2]}pm: {show} ({shows[show][0]}-{shows[show][1]})") # order from dict
		else:
			lines.append(f"{shows[show][2]}pm: {show}")

	return lines

def next_day(start=None, day: int = 0, latest_same_day_hour: int = 20):
    """Get date of next day of week following given date

    Args:
        start: date to search from. Defaults to current time.
        day: day of the week: sun=0, ..., sat=6. Defaults to 0.
        latest_same_day_hour: latest hour to consider next day to be that day 
            i.e. if set to 6, anytime after 6 it will get the following week 
            while 6 or before will get that day. Defaults to 20.

    Returns:
        pendulum.datetime: date with year. month, and day set
    """
    if not start:
        start = pendulum.from_timestamp(time.time(), tz=Resources.timezone_str)
    start = start.subtract(hours=latest_same_day_hour)
    start = start.start_of('day')
    start = start.next(day)
    return start

def extract_schedule(file, start_hour):
	class Time:
		def __init__(self, hour, min):
			self.hour = hour
			self.min = min

		def incr(self, min):
			hours = max(1, min // 60)
			if self.min + min >= 60:
				self.hour += hours
				self.min = self.min + min - (hours*60)
			else:
				self.min += min

		def __str__(self):
			return f"{self.hour}{':'+str(self.min) if self.min else ''}"
	
	wb = load_workbook(filename=file)
	sheet = wb.active
	start_hour = int(sheet.cell(1, 2).value.split(':')[0])
	data = {}
	# go through all rows starting form second (1-based indexing)
	for row in sheet.iter_rows(2):
		# ignore if date isn't 1st column element (0-based indexing here)
		if not isinstance(row[0].value, datetime.datetime):
			continue

		entries = []

		prev = row[1].value
		entry = {
			'title': prev,
			'start': str(start_hour),
			'end': ''
		}
		time = Time(start_hour, 0)
		# get showtime data
		for cell in row[2:]:
			time.incr(10)

			if isinstance(cell, openpyxl.cell.cell.MergedCell) or (cell.value == None and entry['title'] == None):
				continue

			entry['end'] = str(time)
			entries.append(entry)
			entry = {
				'title': cell.value,
				'start': str(time),
				'end': ''
			}
			prev == cell.value


		time.incr(10)
		entry['end'] = str(time)
		entries.append(entry)

		data[str(row[0].value)] = entries
	return data
